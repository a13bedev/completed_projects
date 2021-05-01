"""
Reads all the Excel files in the current working directory
and outputs them as CSV files.
If Excel file contain multiple sheets, program creates one CSV file per sheet.
The filenames of the CSV files would be <excel filename>_<sheet title>.csv.
"""

import openpyxl
import csv
import sys
import os


def convert() -> None:
    for file in os.listdir():
        if not file.endswith('.xlsx'):  # Skip all non.xlsx files
            continue
        wb = openpyxl.load_workbook(file)
        sheets = wb.sheetnames
        # Iterate through sheets of current file
        for i in range(len(sheets)):
            ws = wb[sheets[i]]
            # Create csv file and writer for it.
            with open(os.path.basename(file) + '_' + sheets[i] + '.csv', 'w', newline='') as output_file:
                writer = csv.writer(output_file)
                # Write down a row data lists to csv file.
                for lst in [[cell.value for cell in row] for row in ws.iter_rows()]:
                    writer.writerow(lst)
    print("DONE")


def main():
    path = sys.argv[1]
    # Iterate through directory for .xlsx files
    os.chdir(path)
    convert()


if __name__ == '__main__':
    main()
