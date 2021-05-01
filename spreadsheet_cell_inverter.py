"""
This program
takes file name as command line argument and
inverts row and columns of cells in Excel spreadsheet.
It looks like:
Before:
0  A  B  C  D
1  1  2  3  4
After:
0  1
A  1
B  2
C  3
D  4
"""

# This version saves only values of cells and loses style attributes.

import sys
import openpyxl


def invert(file_name: str) -> None:
    wb = openpyxl.load_workbook(file_name)
    # Get active worksheet.
    ws = wb.active
    # Store title and max number of rows and columns
    original_title = ws.title
    num_of_rows = ws.max_row
    num_of_cols = ws.max_column
    # Create new worksheet.
    new_title = "new sheet"
    new_ws = wb.create_sheet(new_title)
    for i in range(1, num_of_rows + 1):
        for j in range(1, num_of_cols + 1):
            new_ws.cell(row=j, column=i).value = ws.cell(row=i, column=j).value
    # Remove sheet with original title.
    wb.remove(wb[original_title])
    # Save new sheet with original title.
    new_ws.title = original_title
    wb.save(file_name)
    wb.close()


def main():
    if len(sys.argv) > 1:
        f_name = sys.argv[1]
        invert(f_name)
    else:
        print('Argument not specified.')


if __name__ == '__main__':
    main()
