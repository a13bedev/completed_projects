"""
This program takes a number N from command line and
creates a N*N multiplication table in Excel spreadsheet.
It looks like:
     A   B   C   D   etc.
1    1   2   3   4   ...
2    2   4   6   8   ...
3    3   6   9   12  ...
4    4   8   12  16  ...
etc. ... ... ... ... ...
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font


def create_table(n: int) -> None:
    wb = Workbook()
    ws = wb['Sheet']
    # Set row and column for a table and make em bold.
    for i in range(2, n + 2):
        ws.cell(row=1, column=i).value = i-1
        ws.cell(row=1, column=i).font = Font(bold=True)
        ws.cell(row=i, column=1).value = i-1
        ws.cell(row=i, column=1).font = Font(bold=True)
    # Set values for a table.
    for i in range(2, n + 2):
        for j in range(2, n + 2):
            ws.cell(row=i, column=j).value = ws.cell(row=1, column=j).value * \
                                             ws.cell(row=i, column=1).value
    # Save workbook in a file
    wb.save(f'multiTable{n}x{n}.xlsx')
    wb.close()


def main():
    if len(sys.argv) > 1:
        n = int(sys.argv[1])
        create_table(n)
    else:
        print('Argument not specified.')


if __name__ == '__main__':
    main()
